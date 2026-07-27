import asyncio
import hashlib
import logging
import math
from decimal import Decimal
from io import BytesIO, StringIO

from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import RedirectResponse, Response
from pydantic import BaseModel
from sqlalchemy.ext.asyncio import AsyncSession

from app.config.settings import settings
from app.db import get_db
from app.dependencies.auth import get_current_user_id
from app.engine.approval_pdf import OwnerInfo, generate_approval_pdf
from app.engine.boq import QuantityEngine
from app.engine.models import PlotConfig
from app.engine.pdf import render_pdf
from app.engine.plan_geometry import build_floor_drawing
from app.engine.structural_drawing_set import generate_structural_drawing_set
from app.models.project import Project
from app.quality.ccqs import compute_ccqs_deterministic
from app.services.access import get_accessible_project
from app.services.plans import get_effective_plan_tier, tier_at_least
from app.services import (
    layout_store,
    plinth_beam_design,
    structagent_client,
    structural_store,
)
from app.services.plot_config import plot_config_from_project
from app.services.storage import get_storage

logger = logging.getLogger(__name__)
router = APIRouter()

# ReportLab builds the whole document in memory; Cloud Run's filesystem is
# RAM-backed so spooling to /tmp would not help. Bounding concurrency is the
# only thing that actually stops an OOM taking the instance down.
_EXPORT_SEM = asyncio.Semaphore(settings.export_max_concurrency)


def _artifact_key(project_id: str, layout_id: str, ext: str, content: bytes) -> str:
    digest = hashlib.sha256(content).hexdigest()[:16]
    return f"exports/{project_id}/{layout_id}/{digest}.{ext}"


async def _deliver(
    content: bytes, media_type: str, filename: str, key: str
) -> Response:
    """Persist to R2 (best-effort) and return the artifact to the caller."""
    storage = get_storage()
    try:
        await storage.put_bytes(key, content, media_type)
    except Exception:
        logger.warning("R2 upload failed for %s — serving inline", key, exc_info=True)

    if settings.export_delivery_mode == "redirect":
        url = storage.signed_url(key)
        if url:
            return RedirectResponse(url=url, status_code=307)

    return Response(
        content=content,
        media_type=media_type,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


async def _get_plan_tier(user_id: str, db: AsyncSession) -> str:
    return await get_effective_plan_tier(user_id, db)


def _to_float(v) -> float:
    return float(v) if isinstance(v, Decimal) else v


def _cfg_from_project(project: Project) -> PlotConfig:
    return plot_config_from_project(project)


async def _get_project(project_id: str, user_id: str, db: AsyncSession) -> Project:
    return await get_accessible_project(project_id, user_id, db)


async def _maybe_structural_design(
    project_id: str, layout_id: str, geometry: dict, db: AsyncSession
) -> dict | None:
    return await structural_store.design_surface(project_id, layout_id, geometry, db)


# ── PDF export ────────────────────────────────────────────────────────────────


@router.get("/projects/{project_id}/export/pdf")
async def export_pdf(
    project_id: str,
    layout_id: str = "A",
    user_id: str = Depends(get_current_user_id),
    db: AsyncSession = Depends(get_db),
) -> Response:
    project = await _get_project(project_id, user_id, db)
    cfg = _cfg_from_project(project)

    stored = await layout_store.get_or_generate_layouts(project, db)
    row = next((r for r in stored if r.layout_key == layout_id), None)
    if row is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Layout {layout_id!r} not found",
        )
    layout = layout_store.engine_layout_from_geometry(row.geometry)

    design = await _maybe_structural_design(project_id, layout_id, row.geometry, db)
    if design is not None:
        geom = design.get("final_geometry") or row.geometry
        layout = layout_store.engine_layout_from_geometry(geom)

    annotations = getattr(project, "annotations", None) or {}
    async with _EXPORT_SEM:
        pdf_bytes = render_pdf(
            project.name,
            layout,
            cfg,
            project.num_bedrooms,
            annotations=annotations or None,
            structural_design=design,
            watermark_preliminary=True,
        )

    filename = f"planforge-{project_id}-layout-{layout_id}.pdf"
    return await _deliver(
        pdf_bytes,
        "application/pdf",
        filename,
        _artifact_key(project_id, layout_id, "pdf", pdf_bytes),
    )


# ── Structural Drawing Set export ─────────────────────────────────────────────


@router.get("/projects/{project_id}/export/structural-drawing-set")
async def export_structural_drawing_set(
    project_id: str,
    layout_id: str = "A",
    user_id: str = Depends(get_current_user_id),
    db: AsyncSession = Depends(get_db),
) -> Response:
    """Export a 6-page Structural Drawing Set PDF (Column & Footing Plan,
    Footing Details, Plinth Beam Plan, Plinth Beam Details, Roof Beam & Slab
    Plan, Roof Beam Details).

    Gated on approved + designed layout: returns 409 if either approval or
    structural design is missing. Computes plinth beams fresh at export time
    (they depend on wall geometry only, not external structapi data).
    """
    project = await _get_project(project_id, user_id, db)
    cfg = _cfg_from_project(project)

    stored = await layout_store.get_or_generate_layouts(project, db)
    row = next((r for r in stored if r.layout_key == layout_id), None)
    if row is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Layout {layout_id!r} not found",
        )

    # Gate 1: Check approval
    revision = await structural_store.find_revision_for_hash(
        project_id,
        layout_id,
        structural_store.geometry_hash(row.geometry),
        db,
    )
    if revision is None:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail={
                "code": "not_approved",
                "help": (
                    "Approve the architectural plan first: "
                    f"POST /api/projects/{project_id}/structural/approve "
                    f'{{"layout_id": "{layout_id}"}}'
                ),
            },
        )

    # Gate 2: Check structural design exists
    design = await structural_store.design_surface(
        project_id, layout_id, row.geometry, db
    )
    if design is None:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail={
                "code": "not_designed",
                "help": (
                    "Run structural design first: "
                    f"POST /api/projects/{project_id}/structural "
                    f'{{"layout_id": "{layout_id}"}}'
                ),
            },
        )

    # Build layout object (prefer final_geometry if available from design)
    geom = design.get("final_geometry") or row.geometry
    layout = layout_store.engine_layout_from_geometry(geom)

    # Build ground-floor drawing (plinth beams live on GF)
    drawing = build_floor_drawing(layout.ground_floor, cfg)

    # Compute plinth beam design fresh at export time
    try:
        plinth_beams_data = await plinth_beam_design.design_plinth_beams(drawing.walls)
    except structagent_client.StructuralAPIError as exc:
        raise HTTPException(status_code=status.HTTP_502_BAD_GATEWAY, detail=str(exc))

    # Generate the 6-page PDF
    async with _EXPORT_SEM:
        pdf_bytes = generate_structural_drawing_set(
            project_name=project.name,
            cfg=cfg,
            columns=drawing.columns,
            walls=drawing.walls,
            plinth_beams_data=plinth_beams_data,
            structural_design=design,
            floor_plan=layout.first_floor,
            layout=layout,
            num_bedrooms=project.num_bedrooms,
        )

    filename = f"planforge-structural-drawings-{project_id}-layout-{layout_id}.pdf"
    return await _deliver(
        pdf_bytes,
        "application/pdf",
        filename,
        _artifact_key(project_id, layout_id, "pdf", pdf_bytes),
    )


# ── Layout quality ────────────────────────────────────────────────────────────


@router.get("/projects/{project_id}/layouts/{layout_id}/quality")
async def layout_quality(
    project_id: str,
    layout_id: str,
    user_id: str = Depends(get_current_user_id),
    db: AsyncSession = Depends(get_db),
) -> dict:
    project = await _get_project(project_id, user_id, db)
    cfg = _cfg_from_project(project)

    stored = await layout_store.get_or_generate_layouts(project, db)
    row = next((r for r in stored if r.layout_key == layout_id), None)
    if row is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Layout {layout_id!r} not found",
        )
    layout = layout_store.engine_layout_from_geometry(row.geometry)

    annotations = getattr(project, "annotations", None) or {}
    pdf_bytes = render_pdf(
        project.name, layout, cfg, project.num_bedrooms, annotations=annotations or None
    )
    return compute_ccqs_deterministic(pdf_bytes).as_dict()


# ── Approval PDF export ───────────────────────────────────────────────────────


class ApprovalPdfRequest(BaseModel):
    owner_name: str
    survey_number: str
    locality: str
    engineer_name: str
    license_number: str
    municipality: str | None = None


@router.post("/projects/{project_id}/export/approval-pdf")
async def export_approval_pdf(
    project_id: str,
    body: ApprovalPdfRequest,
    layout_id: str = "A",
    user_id: str = Depends(get_current_user_id),
    db: AsyncSession = Depends(get_db),
) -> Response:
    project = await _get_project(project_id, user_id, db)
    cfg = _cfg_from_project(project)

    stored = await layout_store.get_or_generate_layouts(project, db)
    row = next((r for r in stored if r.layout_key == layout_id), None)
    if row is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Layout {layout_id!r} not found",
        )
    layout = layout_store.engine_layout_from_geometry(row.geometry)

    municipality = body.municipality or getattr(project, "municipality", None) or ""
    owner = OwnerInfo(
        owner_name=body.owner_name,
        survey_number=body.survey_number,
        locality=body.locality,
        engineer_name=body.engineer_name,
        license_number=body.license_number,
        municipality=municipality,
    )

    async with _EXPORT_SEM:
        pdf_bytes = generate_approval_pdf(layout, cfg, owner, layout_id)

    filename = f"planforge-approval-{project_id}-layout-{layout_id}.pdf"
    return await _deliver(
        pdf_bytes,
        "application/pdf",
        filename,
        _artifact_key(project_id, layout_id, "pdf", pdf_bytes),
    )


# ── DXF export ────────────────────────────────────────────────────────────────


@router.get("/projects/{project_id}/export/dxf")
async def export_dxf(
    project_id: str,
    layout_id: str = "A",
    user_id: str = Depends(get_current_user_id),
    db: AsyncSession = Depends(get_db),
) -> Response:
    plan = await _get_plan_tier(user_id, db)
    if not tier_at_least(plan, "basic"):
        raise HTTPException(
            status_code=402, detail="DXF export requires Basic or Pro plan."
        )

    try:
        import ezdxf  # noqa: F401
    except ImportError:
        raise HTTPException(
            status_code=status.HTTP_501_NOT_IMPLEMENTED,
            detail="ezdxf not installed. Run: uv add ezdxf",
        )

    project = await _get_project(project_id, user_id, db)
    cfg = _cfg_from_project(project)

    stored = await layout_store.get_or_generate_layouts(project, db)
    row = next((r for r in stored if r.layout_key == layout_id), None)
    if row is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Layout {layout_id!r} not found",
        )
    layout = layout_store.engine_layout_from_geometry(row.geometry)

    design = await _maybe_structural_design(project_id, layout_id, row.geometry, db)
    if design is not None:
        geom = design.get("final_geometry") or row.geometry
        layout = layout_store.engine_layout_from_geometry(geom)

    async with _EXPORT_SEM:
        dxf_bytes = _render_dxf(project.name, layout, cfg, structural_design=design)

    filename = f"planforge-{project_id}-layout-{layout_id}.dxf"
    return await _deliver(
        dxf_bytes,
        "application/dxf",
        filename,
        _artifact_key(project_id, layout_id, "dxf", dxf_bytes),
    )


# Calibrated to the same reference print scale as plan_geometry's LabelBox
# font_pt sizing (1 pt at 1:100) so DXF label heights match the PDF projection.
_DXF_PT_TO_M = 0.000352778 * 100


def _draw_stair_dxf(msp, stair, layer: str, z: float) -> None:
    """Treads + mid-flight break line + UP arrow, from canonical StairGeometry."""
    for x1, y1, x2, y2 in stair.treads:
        msp.add_line(
            (x1, y1, z), (x2, y2, z), dxfattribs={"layer": layer, "lineweight": 25}
        )

    bx1, by1, bx2, by2 = stair.break_line
    msp.add_line(
        (bx1, by1, z),
        (bx2, by2, z),
        dxfattribs={"layer": layer, "lineweight": 25, "linetype": "DASHED"},
    )

    ax1, ay1, ax2, ay2 = stair.arrow
    msp.add_line(
        (ax1, ay1, z), (ax2, ay2, z), dxfattribs={"layer": layer, "lineweight": 25}
    )
    ang = math.atan2(ay2 - ay1, ax2 - ax1)
    for da in (2.6, -2.6):
        msp.add_line(
            (ax2, ay2, z),
            (
                ax2 - 0.15 * math.cos(ang + da * 0.5),
                ay2 - 0.15 * math.sin(ang + da * 0.5),
                z,
            ),
            dxfattribs={"layer": layer, "lineweight": 25},
        )
    ux, uy = stair.up_label_xy
    msp.add_mtext(
        "UP",
        dxfattribs={
            "layer": layer,
            "char_height": 0.18,
            "insert": (ux, uy, z),
            "attachment_point": 5,
        },
    )


def _draw_labels_dxf(msp, labels, layer: str, z: float) -> None:
    """Room name + size labels, from canonical LabelBox — never truncated;
    plain TEXT per line (no MTEXT \\P codes), rotated/leadered per the fit."""
    from ezdxf.enums import TextEntityAlignment

    for lb in labels:
        if lb.leader is not None:
            tx, ty = lb.leader
            msp.add_line(
                (lb.cx, lb.cy, z),
                (tx, ty, z),
                dxfattribs={"layer": layer, "lineweight": 18},
            )
            msp.add_circle((tx, ty, z), radius=0.03, dxfattribs={"layer": layer})

        line_h = lb.font_pt * _DXF_PT_TO_M * 1.25
        top = (len(lb.lines) - 1) * line_h / 2
        rotation = 90.0 if lb.rotated else 0.0
        for i, text in enumerate(lb.lines):
            along = top - i * line_h
            px, py = (lb.cx - along, lb.cy) if lb.rotated else (lb.cx, lb.cy + along)
            height = lb.font_pt * _DXF_PT_TO_M * (1.0 if i == 0 else 0.9)
            entity = msp.add_text(
                text,
                dxfattribs={"layer": layer, "height": height, "rotation": rotation},
            )
            entity.set_placement((px, py, z), align=TextEntityAlignment.MIDDLE_CENTER)


def _draw_dim_chains_dxf(msp, drawing, layer: str) -> None:
    """Multi-level dimension chains (room + plot/setback, all 4 sides) from
    the canonical FloorDrawing — entry text is pre-formatted ft-in."""
    for chain in drawing.dim_chains:
        if not chain.entries:
            continue
        horizontal = chain.side in ("bottom", "top")
        for entry in chain.entries:
            if horizontal:
                base = p1 = (entry.start, chain.coord)
                p2 = (entry.end, chain.coord)
                angle = 0
            else:
                base = p1 = (chain.coord, entry.start)
                p2 = (chain.coord, entry.end)
                angle = 90
            try:
                dim = msp.add_linear_dim(
                    base=base,
                    p1=p1,
                    p2=p2,
                    angle=angle,
                    dimstyle="ARCH_MM",
                    dxfattribs={"layer": layer, "lineweight": 18},
                )
                dim.set_text(entry.text)
                dim.render()
            except Exception as exc:
                logger.warning("Dimension render failed: %s", exc)


def _render_dxf(
    project_name: str, layout, cfg: PlotConfig, structural_design: dict | None = None
) -> bytes:
    import ezdxf
    from ezdxf import colors

    from app.engine.cad_advanced import (
        draw_building_footprint,
        draw_compound_wall,
        draw_furniture,
        draw_open_terrace,
        draw_setback_zones,
        draw_sized_columns,
        draw_structural_grid,
        draw_structural_schedule,
        shapely_poly_to_dxf,
        solid_fill_polygon,
    )
    from app.engine.cad_blocks import (
        define_opening_blocks,
        insert_door,
        insert_ventilator,
        insert_window,
    )
    from app.engine.cad_primitives import (
        draw_north_arrow,
        draw_scale_bar,
        draw_title_block,
    )
    from app.engine.plan_geometry import (
        build_floor_drawing,
        opening_boxes,
        wall_polygons,
    )

    design_data: dict = {}
    if structural_design and structural_design.get("status") not in (None, "stale"):
        design_data = (structural_design.get("structapi") or {}).get("data") or {}

    doc = ezdxf.new("R2010", setup=True)  # setup=True loads standard linetypes
    doc.header["$INSUNITS"] = 6  # metres (geometry stored in metres)
    doc.header["$MEASUREMENT"] = 1  # metric hatch/linetype scaling
    doc.header["$LWDISPLAY"] = 1  # show lineweights in DXF viewers

    layer_defs = [
        ("PLOT-BOUNDARY", colors.GREEN, 0.25),
        ("A-WALL-BRICK", colors.RED, 0.50),
        ("A-WALL-INT", colors.YELLOW, 0.35),
        ("A-DOOR", colors.CYAN, 0.25),
        ("A-WINDOW", colors.BLUE, 0.25),
        ("A-STAIR", colors.WHITE, 0.25),
        ("A-VENTILATOR", colors.MAGENTA, 0.18),
        ("A-TITLE", colors.WHITE, 0.50),
        ("S-COLUMN", colors.WHITE, 0.35),
        ("S-BEAM", colors.WHITE, 0.35),
        ("S-GRID", colors.GRAY, 0.18),
        ("S-COLUMNS-SIZED", colors.RED, 0.50),
        ("S-SCHEDULE", colors.WHITE, 0.18),
        ("DIM-LINE", colors.GRAY, 0.18),
        ("TEXT", colors.WHITE, 0.18),
        # Advanced CAD layers
        ("A-FOOTPRINT", colors.WHITE, 0.70),
        ("A-COMPOUND-WALL", colors.GREEN, 0.35),
        ("A-TERRACE", colors.CYAN, 0.18),
        ("A-FURNITURE", colors.BLUE, 0.18),
        ("DIM-SETBACK", colors.GRAY, 0.18),
    ]
    # Structural layers are frozen by default so architectural drawing stays clean
    structural_layers = {"S-COLUMN", "S-BEAM", "S-GRID"}

    for lname, color, lw in layer_defs:
        lyr = doc.layers.new(lname)
        lyr.color = color
        lyr.lineweight = int(lw * 100)
        if lname in structural_layers:
            lyr.freeze()

    # DEFPOINTS — non-printing layer required by DXF spec for dimension attachment points
    if "DEFPOINTS" not in doc.layers:
        _dp = doc.layers.new("DEFPOINTS")
        _dp.dxf.plot = 0  # non-printing

    define_opening_blocks(doc)

    # Architectural dimension style — created once per document
    # ezdxf is imported at the top of this function; use it directly
    _ds = doc.dimstyles.new("ARCH_MM")
    _ds.dxf.dimtxt = 0.25  # text height (m) → 2.5mm on paper at 1:100
    _ds.dxf.dimasz = 0.15  # arrow size
    _ds.dxf.dimtad = 1  # text above dim line
    _ds.dxf.dimexo = 0.10  # extension line offset
    _ds.dxf.dimexe = 0.15  # extension line overshoot
    _ds.dxf.dimgap = 0.08  # gap between text and dim line
    _ds.dxf.dimdec = 0  # no decimal places (text overridden by set_text to ft-in)
    try:
        _ds.set_arrows(blk=ezdxf.ARROWS.architectural_tick)
    except Exception as exc:
        logger.warning("Archtick arrow unavailable: %s", exc)

    # Register DASHED linetype (used by plot boundary and structural grid)
    if "DASHED" not in doc.linetypes:
        doc.linetypes.new("DASHED", dxfattribs={"description": "Dashed _ _ _"})

    msp = doc.modelspace()

    # ── Plot boundary ─────────────────────────────────────────────────────────
    if (
        cfg.plot_shape == "quadrilateral"
        and cfg.plot_corners
        and len(cfg.plot_corners) == 4
    ):
        boundary_pts = [(float(x), float(y)) for x, y in cfg.plot_corners]
    else:
        boundary_pts = [
            (0.0, 0.0),
            (cfg.plot_width, 0.0),
            (cfg.plot_width, cfg.plot_length),
            (0.0, cfg.plot_length),
        ]
    msp.add_lwpolyline(
        boundary_pts,
        close=True,
        dxfattribs={"layer": "PLOT-BOUNDARY", "linetype": "DASHED"},
    )

    # Collect all floor plans (including optional second/basement floors)
    floor_plans = [layout.ground_floor, layout.first_floor]
    if layout.second_floor:
        floor_plans.append(layout.second_floor)
    if layout.basement_floor:
        floor_plans.append(layout.basement_floor)

    global_min_x = global_min_y = float("inf")
    global_max_x = global_max_y = float("-inf")

    # Ground-floor building extents needed for post-loop setback callouts
    gf_bld_x = gf_bld_y = gf_bld_w = gf_bld_d = 0.0
    gf_main_door_x: float | None = None

    for floor_plan in floor_plans:
        z_offset = float(floor_plan.floor) * 3.0
        rooms = floor_plan.rooms
        if not rooms:
            continue

        bld_x = min(r.x for r in rooms)
        bld_y = min(r.y for r in rooms)
        bld_w = max(r.x + r.width for r in rooms) - bld_x
        bld_d = max(r.y + r.depth for r in rooms) - bld_y

        if floor_plan.floor == 0:
            gf_bld_x, gf_bld_y, gf_bld_w, gf_bld_d = bld_x, bld_y, bld_w, bld_d

        global_min_x = min(global_min_x, bld_x)
        global_min_y = min(global_min_y, bld_y)
        global_max_x = max(global_max_x, bld_x + bld_w)
        global_max_y = max(global_max_y, bld_y + bld_d)

        # Canonical drawing for this floor — single source of truth for
        # walls/openings/columns/dims/labels/stair (Sprint 4/5).
        drawing = build_floor_drawing(floor_plan, cfg)

        if floor_plan.floor == 0:
            gf_main_door_x = next(
                (o.cx for o in drawing.openings if getattr(o, "is_main", False)), None
            )

        # 1-2. Walls: poché fill from the union of wall footprints with
        # opening boxes already subtracted (IS:962/AIA convention).
        wall_polys = wall_polygons(
            drawing.walls, openings=opening_boxes(drawing.openings)
        )
        for kind, lyr_name in (
            ("external", "A-WALL-BRICK"),
            ("internal", "A-WALL-INT"),
        ):
            geom = wall_polys[kind]
            if geom.is_empty:
                continue
            shapely_poly_to_dxf(msp, geom, layer=lyr_name, z=z_offset)
            solid_fill_polygon(msp, geom, layer=lyr_name, z=z_offset)

        # 3. Opening block inserts (PF_DOOR/PF_WINDOW/PF_VENT — Task 14),
        # positioned from the canonical Opening (hinge/swing-aware for doors).
        for op in drawing.openings:
            if op.kind == "door":
                jamb_x = 2 * op.cx - op.hinge_x
                jamb_y = 2 * op.cy - op.hinge_y
                rotation = math.degrees(
                    math.atan2(jamb_y - op.hinge_y, jamb_x - op.hinge_x)
                )
                insert_door(
                    msp,
                    op.hinge_x,
                    op.hinge_y,
                    rotation,
                    z=z_offset,
                    mirror=op.swing_cw,
                )
                if getattr(op, "is_main", False):
                    off = op.wall_thickness / 2 + 0.15
                    tx, ty = (
                        (op.cx, op.cy - off)
                        if op.is_horizontal
                        else (
                            op.cx - off,
                            op.cy,
                        )
                    )
                    msp.add_mtext(
                        "MD",
                        dxfattribs={
                            "layer": "TEXT",
                            "char_height": 0.15,
                            "insert": (tx, ty, z_offset),
                            "attachment_point": 5,
                        },
                    )
                continue
            if op.is_horizontal:
                ins_x, ins_y, rotation = op.cx - op.width / 2, op.cy, 0.0
            else:
                ins_x, ins_y, rotation = op.cx, op.cy - op.width / 2, 90.0
            insert_fn = insert_window if op.kind == "window" else insert_ventilator
            insert_fn(msp, ins_x, ins_y, rotation, z=z_offset)

        # 4. Staircase treads + cut line + UP arrow
        if drawing.stair is not None:
            _draw_stair_dxf(msp, drawing.stair, layer="A-STAIR", z=z_offset)

        # 5. Room labels (never truncated — fit ladder from S4.3)
        _draw_labels_dxf(msp, drawing.labels, layer="TEXT", z=z_offset)

        # 6. Columns (already deduped by junction derivation)
        for col in drawing.columns:
            half = col.size / 2
            pts_col = [
                (col.cx - half, col.cy - half),
                (col.cx + half, col.cy - half),
                (col.cx + half, col.cy + half),
                (col.cx - half, col.cy + half),
            ]
            msp.add_lwpolyline(
                pts_col,
                close=True,
                dxfattribs={"layer": "S-COLUMN", "elevation": z_offset},
            )

        # 6a. Bold building outline (unary_union of room boxes)
        footprint = draw_building_footprint(msp, rooms, layer="A-FOOTPRINT", z=z_offset)

        # 6b. Structural grid (ground floor only)
        if floor_plan.floor == 0:
            draw_structural_grid(
                msp, rooms, bld_x, bld_y, bld_w, bld_d, layer="S-GRID", z=z_offset
            )
            if design_data:
                draw_sized_columns(
                    msp,
                    drawing.columns,
                    drawing.walls,
                    design_data.get("columns") or {},
                    layer="S-COLUMNS-SIZED",
                    z=z_offset,
                )

        # 6c. Furniture per room
        for room in rooms:
            draw_furniture(msp, room, layer="A-FURNITURE", z=z_offset)

        # 6d. Open terrace hatching (ground floor only)
        if floor_plan.floor == 0 and footprint is not None:
            from shapely.geometry import Polygon as _SPoly
            from shapely.geometry import box as _sbox

            plot_poly = (
                _SPoly([(float(x), float(y)) for x, y in cfg.plot_corners])
                if cfg.plot_shape == "quadrilateral" and cfg.plot_corners
                else _sbox(0, 0, cfg.plot_width, cfg.plot_length)
            )
            draw_open_terrace(msp, plot_poly, footprint, layer="A-TERRACE", z=z_offset)

        # 7. Dimension chains (room + plot/setback levels, all 4 sides)
        _draw_dim_chains_dxf(msp, drawing, layer="DIM-LINE")

    # ── North arrow (top-right, drawn once outside floor loop) ───────────────
    if global_max_x < float("inf"):
        north_dir = getattr(cfg, "road_side", "S") or "S"
        draw_north_arrow(
            msp,
            cx=global_max_x + 2.5,
            cy=global_max_y - 1.5,
            north_dir=north_dir,
            size=0.8,
            layer="TEXT",
        )
        draw_scale_bar(
            msp, x=global_max_x + 2.0, y=global_max_y - 3.5, layer="TEXT", z=0.0
        )

        # ── Setback dimension callouts (ground floor extents) ─────────────────
        draw_setback_zones(
            msp, cfg, gf_bld_x, gf_bld_y, gf_bld_w, gf_bld_d, layer="DIM-SETBACK", z=0.0
        )

        # ── Compound boundary wall with gate ─────────────────────────────────
        draw_compound_wall(
            msp, cfg, layer="A-COMPOUND-WALL", z=0.0, gate_cx=gf_main_door_x
        )

        # ── Title block (below the drawing) ───────────────────────────────────
        gf_sqft = sum(r.area for r in layout.ground_floor.rooms) * 10.764
        ff_sqft = sum(r.area for r in layout.first_floor.rooms) * 10.764
        draw_title_block(
            msp,
            project_name=project_name,
            layout_id=layout.id,
            gf_area_sqft=gf_sqft,
            ff_area_sqft=ff_sqft,
            plot_w=cfg.plot_width,
            plot_l=cfg.plot_length,
            insert_x=global_min_x,
            insert_y=global_min_y - 5.5,
        )

        if design_data:
            draw_structural_schedule(
                msp,
                design_data.get("columns") or {},
                design_data.get("beams") or {},
                insert_x=global_max_x + 2.0,
                insert_y=global_min_y - 5.5,
                layer="S-SCHEDULE",
                z=0.0,
            )

    # ezdxf writes DXF as text (not binary) — use StringIO then encode
    text_buf = StringIO()
    doc.write(text_buf)
    return text_buf.getvalue().encode("utf-8")


# ── BOQ export ────────────────────────────────────────────────────────────────


@router.get("/projects/{project_id}/boq")
async def export_boq(
    project_id: str,
    layout_id: str = "A",
    fmt: str = "json",
    city: str = "Generic",
    user_id: str = Depends(get_current_user_id),
    db: AsyncSession = Depends(get_db),
) -> Response:
    project = await _get_project(project_id, user_id, db)
    cfg = _cfg_from_project(project)

    stored = await layout_store.get_or_generate_layouts(project, db)
    row = next((r for r in stored if r.layout_key == layout_id), None)
    if row is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Layout {layout_id!r} not found",
        )
    layout = layout_store.engine_layout_from_geometry(row.geometry)

    design = await _maybe_structural_design(project_id, layout_id, row.geometry, db)
    structural_quantities = None
    if design is not None:
        geom = design.get("final_geometry") or row.geometry
        layout = layout_store.engine_layout_from_geometry(geom)
        structural_quantities = ((design.get("structapi") or {}).get("data") or {}).get(
            "quantities"
        )

    engine = QuantityEngine()
    boq = engine.calculate(
        layout,
        cfg,
        project_name=project.name,
        city=city,
        structural_quantities=structural_quantities,
    )

    if fmt == "excel":
        plan = await _get_plan_tier(user_id, db)
        if not tier_at_least(plan, "pro"):
            raise HTTPException(
                status_code=402, detail="BOQ Excel export requires Pro plan."
            )
        # The workbook build is synchronous and allocates the whole file in
        # memory, so it shares the render budget with the PDF/DXF paths.
        async with _EXPORT_SEM:
            return await _boq_excel_response(boq, project_id, layout_id)

    import json

    return Response(
        content=json.dumps(boq.to_dict(), indent=2),
        media_type="application/json",
    )


async def _boq_excel_response(boq, project_id: str, layout_id: str) -> Response:
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        raise HTTPException(
            status_code=status.HTTP_501_NOT_IMPLEMENTED,
            detail="openpyxl not installed. Run: uv add openpyxl",
        )

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "BOQ"

    # Title row 1: project + layout
    ws.merge_cells("A1:F1")
    ws["A1"] = f"Bill of Quantities — {boq.project_name} / Layout {boq.layout_id}"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A1"].alignment = Alignment(horizontal="center")

    # Title row 2: city / rates note
    ws.merge_cells("A2:F2")
    ws["A2"] = boq.rates_note
    ws["A2"].font = Font(italic=True, size=10, color="555555")
    ws["A2"].alignment = Alignment(horizontal="center")

    # Column headers
    headers = [
        "S.No",
        "Item Description",
        "Quantity",
        "Unit",
        "Rate (₹)",
        "Amount (₹)",
        "Basis",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1E3A5F")
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, item in enumerate(boq.line_items, start=5):
        ws.cell(row=row_idx, column=1, value=item.item)
        ws.cell(row=row_idx, column=2, value=item.description)
        ws.cell(row=row_idx, column=3, value=item.quantity)
        ws.cell(row=row_idx, column=4, value=item.unit)
        ws.cell(row=row_idx, column=5, value=round(item.rate, 2) if item.rate else "")
        ws.cell(row=row_idx, column=6, value=round(item.amount) if item.amount else "")
        ws.cell(row=row_idx, column=7, value=item.basis)

    # Total row
    total_row = len(boq.line_items) + 6
    ws.cell(row=total_row, column=2, value="TOTAL ESTIMATED COST")
    ws.cell(row=total_row, column=2).font = Font(bold=True)
    ws.cell(row=total_row, column=6, value=round(boq.total_cost))
    ws.cell(row=total_row, column=6).font = Font(bold=True)

    # City comparison row
    if boq.city != "Generic" and boq.cost_difference is not None:
        diff = boq.cost_difference
        diff_label = (
            f"vs Generic: +₹{diff:,.0f} more"
            if diff > 0
            else f"vs Generic: ₹{abs(diff):,.0f} less"
        )
        compare_row = total_row + 1
        ws.cell(row=compare_row, column=2, value=diff_label)
        ws.cell(row=compare_row, column=2).font = Font(
            italic=True, color="CC0000" if diff > 0 else "007700"
        )

    # Auto-width columns (skip merged cells)
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        cell = col[0]
        if hasattr(cell, "column_letter"):
            ws.column_dimensions[cell.column_letter].width = max(max_len + 2, 12)

    # Footer note
    note_row = len(boq.line_items) + 9
    ws.cell(
        row=note_row,
        column=1,
        value="Note: Quantities are approximate. "
        "Verify with site measurements before procurement.",
    )
    ws.cell(row=note_row, column=1).font = Font(italic=True, color="888888")
    ws.cell(row=note_row + 1, column=1, value="Generated by PlanForge")

    buf = BytesIO()
    wb.save(buf)
    xlsx_bytes = buf.getvalue()

    filename = f"planforge-boq-{project_id}-{layout_id}.xlsx"
    return await _deliver(
        xlsx_bytes,
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename,
        _artifact_key(project_id, layout_id, "xlsx", xlsx_bytes),
    )
